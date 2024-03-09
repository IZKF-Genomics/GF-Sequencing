import openpyxl
import xlrd
import os
import difflib
import logging
import Levenshtein
import re
import warnings

from datetime import datetime, timedelta

from utils.utilities import read_cache, write_cache
from config import APPLICATION_MAPPING, SCIEBO_FOLDER_PATH


# Suppress specific openpyxl warnings
warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")
warnings.filterwarnings("ignore", message="Unknown extension is not supported and will be removed")

logger = logging.getLogger(__name__)


def parse_sciebo_report(df, fastq_folder_name):
    sciebo_report_path = find_corresponding_sciebo(fastq_folder_name)
    if sciebo_report_path != None and sciebo_report_path.lower().endswith(".xlsx"):
        parse_sciebo_xlsx_report(df, fastq_folder_name, sciebo_report_path)
    elif sciebo_report_path != None and sciebo_report_path.lower().endswith(".xls"):
        parse_sciebo_xls_report(df, fastq_folder_name, sciebo_report_path)
    else:
        df.loc[fastq_folder_name,"Sciebo Found"] = False
        logger.error("Unsupported file format for sciebo_report")

def parse_sciebo_xls_report(df, fastq_folder_name, report_path):
    """ Gather:
    - Project Name
    - Sequencing Kit
    - Cycles Read 1/2
    - Cycles Index 1/2
    - Density
    - Clusters PF
    - Yield
    - % >= Q30
    """
    sequencing_kit = None
    cycles_read_1 = None
    cycles_index_1 = None
    cycles_read_2 = None
    cycles_index_2 = None
    density = None
    clusters_pf = None
    yields = None
    q_30 = None
    project_name = None
    protocol_name = None
    application = None
    phix_input = None

    file_name = os.path.splitext(os.path.basename(report_path))[0]
    parts = file_name.split('_')
    # Extract the required parts
    date_part = parts[0]  # YYMMDD
    # Convert the date parts to integers for comparison
    year = int(date_part[:2]) + 2000
    month = int(date_part[2:4])
    day = int(date_part[4:6])

    application_part = parts[-1]  # Application
    # Combine the required parts to form the desired string
    protocol_name = f"{date_part}_{application_part}"

    # Find the closest match using difflib
    application = get_application_from_filename(application_part.lower())


    # Open the Excel file
    workbook = xlrd.open_workbook(report_path)

    # Iterate through sheets
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)

        # Iterate through rows and columns
        for i in range(sheet.nrows):
            for j in range(sheet.ncols):
                cell_value = sheet.cell_value(i, j)

                if cell_value is None:
                    continue
                if cell_value == "Cycles Read 1":
                    cycles_read_1 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Index 1":
                    cycles_index_1 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Index 2":
                    cycles_index_2 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Read 2":
                    cycles_read_2 = sheet.cell_value(i, j + 1)
                elif cell_value == "Density":
                    density = sheet.cell_value(i, j + 1)
                elif cell_value == "Clusters PF":
                    clusters_pf = sheet.cell_value(i, j + 1)
                elif cell_value == "Yield":
                    yields = sheet.cell_value(i, j + 1)
                    if type(yields) is str:
                        yields = yields.replace(',', '.')
                        yields =''.join(char for char in yields if char.isdigit() or char == '.')
                elif cell_value == "% >= Q30":
                    q_30 = sheet.cell_value(i, j + 1)
                    if q_30 is None:
                        continue
                    try:
                        if type(q_30) is str:
                            q_30 = q_30.replace('%', '')
                            q_30 = float(q_30.replace(' ',''))
                        if q_30 < 1:
                            q_30 *= 100
                    except BaseException:
                        logger.info(f"q_30 = {q_30} was not handled successfully")
                elif isinstance(cell_value, str) and "kit" in cell_value.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = cell_value.find(':')
                    if index_colon != -1:
                        sequencing_kit = cell_value[index_colon + 1:].strip()
                        sequencing_kit = sequencing_kit.lower()
                elif isinstance(cell_value, str) and "project" in cell_value.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = cell_value.find(':')
                    if index_colon != -1:
                        project_name = cell_value[index_colon + 1:].strip()
                elif isinstance(cell_value, str) and "phix" in cell_value.lower():
                    if (year, month, day) > (2024, 1, 16):
                        # The date is past 16.01.24 and the phix protocol is in place
                        phix_input = sheet.cell_value(i, j + 1)
                    else:
                        # No Protocol prior to the 16.01.24
                        phix_input = None

    df.loc[fastq_folder_name, ['Sequencing Kit', 'Cycles Read 1', 'Cycles Index 1', 'Cycles Read 2', 'Cycles Index 2', 'Density', 'Clusters PF', 'Yields', 'Q 30', 'Name', 'Protocol Name', 'Application', 'Phix Input']] = [sequencing_kit, cycles_read_1, cycles_index_1, cycles_read_2, cycles_index_2, density, clusters_pf, yields, q_30, project_name, protocol_name, application, phix_input]
    df.loc[fastq_folder_name,"Sciebo Found"] = True

def parse_sciebo_xlsx_report(df, fastq_folder_name, report_path):
    """ Gather:
    - Project Name
    - Sequencing Kit
    - Cycles Read 1/2
    - Cycles Index 1/2
    - Density
    - Clusters PF
    - Yield
    - % >= Q30
    """
    sequencing_kit = None
    cycles_read_1 = None
    cycles_index_1 = None
    cycles_read_2 = None
    cycles_index_2 = None
    density = None
    clusters_pf = None
    yields = None
    q_30 = None
    project_name = None
    protocol_name = None
    application = None
    phix_input = None

    file_name = os.path.splitext(os.path.basename(report_path))[0]
    parts = file_name.split('_')
    # Extract the required parts
    date_part = parts[0]  # YYMMDD
    # Convert the date parts to integers for comparison
    year = int(date_part[:2]) + 2000
    month = int(date_part[2:4])
    day = int(date_part[4:6])

    application_part = parts[-1]  # Application
    # Combine the required parts to form the desired string
    protocol_name = f"{date_part}_{application_part}"
    
    # Find the closest match using difflib
    application = get_application_from_filename(application_part.lower())

    wb = openpyxl.load_workbook(filename=report_path, data_only=True)

    # Get All Sheets
    for excel_sheet_name in wb.sheetnames:

        # Get Sheet Object by names
        excel_sheet = wb[excel_sheet_name]

        for i in range(1,excel_sheet.max_row):
            for j in range(1, excel_sheet.max_column):
                excel_sheet_name_cell = excel_sheet.cell(row=i, column=j).value
                if excel_sheet_name_cell == None:
                    continue
                elif excel_sheet_name_cell == "Cycles Read 1":
                    cycles_read_1 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Index 1":
                    cycles_index_1 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Index 2":
                    cycles_index_2 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Read 2":
                    cycles_read_2 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Density":
                    density = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Clusters PF":
                    clusters_pf = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Yield":
                    yields = excel_sheet.cell(row=i, column=j+1).value
                    if type(yields) is str:
                        yields = yields.replace(',', '.')
                        yields =''.join(char for char in yields if char.isdigit() or char == '.')
                elif excel_sheet_name_cell == "% >= Q30":
                    q_30 = excel_sheet.cell(row=i, column=j+1).value
                    if q_30 is None:
                        continue
                    try:
                        if type(q_30) is str:
                            q_30 = q_30.replace('%', '')
                            q_30 = float(q_30.replace(' ',''))
                        if q_30 < 1:
                            q_30 *= 100
                        q_30 = str(q_30)
                    except BaseException:
                        logger.info(f"q_30 = {q_30} was not handled successfully")
                elif type(excel_sheet_name_cell) == str and "phix" in excel_sheet_name_cell.lower():
                    if (year, month, day) > (2024, 1, 16):
                        # The date is past 16.01.24 and the phix protocol is in place
                        phix_input = excel_sheet.cell(row=i, column=j+1).value
                    else:
                        # No Protocol prior to the 16.01.24
                        phix_input = None
                elif type(excel_sheet_name_cell) == str and "kit" in excel_sheet_name_cell.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = excel_sheet_name_cell.find(':')
                    if index_colon != -1:
                        sequencing_kit = excel_sheet_name_cell[index_colon + 1:].strip()
                        sequencing_kit = sequencing_kit.lower()
                elif type(excel_sheet_name_cell) == str and "project" in excel_sheet_name_cell.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = excel_sheet_name_cell.find(':')
                    if index_colon != -1:
                        project_name = excel_sheet_name_cell[index_colon + 1:].strip()

    df.loc[fastq_folder_name, ['Sequencing Kit', 'Cycles Read 1', 'Cycles Index 1', 'Cycles Read 2', 'Cycles Index 2', 'Density', 'Clusters PF', 'Yields', 'Q 30', 'Name', 'Protocol Name', 'Application', 'Phix Input']] = [sequencing_kit, cycles_read_1, cycles_index_1, cycles_read_2, cycles_index_2, density, clusters_pf, yields, q_30, project_name, protocol_name, application, phix_input]
    df.loc[fastq_folder_name,"Sciebo Found"] = True

def get_application_from_filename(filename):
    parts = filename.split('_')
    application_part = parts[-1]  # Last part of the filename
    closest_match = difflib.get_close_matches(application_part.lower(), APPLICATION_MAPPING.keys(), n=1, cutoff=0.6)
    application = APPLICATION_MAPPING.get(closest_match[0], 'unknown') if closest_match else 'unknown'
    return application

###############################################################################
#-------------------- Functions for 'Fastq-Sciebo' mathcing ------------------#
###############################################################################

def find_corresponding_sciebo(fastq_folder):
    cache = read_cache()

    # Check the cache first
    if fastq_folder in cache:
        if cache[fastq_folder] != '':
            logger.info(f"Using cached sciebo file for {fastq_folder}: {cache[fastq_folder]}")
            return cache[fastq_folder]
        else:
            # Check if more than 3 months have passed
            if has_three_months_passed(fastq_folder):
                logger.info("3 months have passed since the given date - Skipping Sciebo search")
                return None

    # Beggining the search for the corresponding sciebo
    sequence_date_prefix = fastq_folder.split('_')[0]
    flowcell_id = fastq_folder.split('_')[3][1:]
    sciebo_files = []

    # Walk through the directory tree
    for folder_path, _, files in os.walk(SCIEBO_FOLDER_PATH):
        for file_name in files:
            sciebo_file_path = os.path.join(folder_path, file_name)
            sciebo_files.append(sciebo_file_path)

    sciebo_candidates = [file for file in sciebo_files if sciebo_date_match(file, sequence_date_prefix)]
    if len(sciebo_candidates) == 0:
        logger.info(f"zero sciebo candidates for {fastq_folder}")
        return None
    elif len(sciebo_candidates) == 1:
        cache[fastq_folder] = sciebo_candidates[0]
        write_cache(cache)
        logger.info(f"single sciebo match found for {fastq_folder}! {sciebo_candidates[0]}")
        return sciebo_candidates[0]
    else:
        logger.info(f"multiple candidates for {fastq_folder}! {sciebo_candidates}")
        for sciebo_file in sciebo_candidates:
            if sciebo_fastq_match(flowcell_id, sciebo_file):
                logger.info(f"sciebo match found from multiple candidates for {fastq_folder}! {sciebo_file}")
                cache[fastq_folder] = sciebo_file
                write_cache(cache)
                return sciebo_file
    logger.info(f"No sciebo match was found for {fastq_folder}")
    cache[fastq_folder] = ''
    write_cache(cache)
    return None

def sciebo_date_match(file_name, desired_date):
    """ Find if the sciebo corresponds to the specific run date
    """
    report_path = file_name
    if file_name.lower().endswith(".xls"):
        # Open the Excel file
        try:
            workbook = xlrd.open_workbook(report_path)
        except:
            return False
        # Iterate through sheets
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            # Iterate through rows and columns
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(i, j)
                    if type(cell_value) is str:
                        if cell_value == "Run name":
                            try:
                                run_name = sheet.cell_value(i, j + 1)
                            except:
                                continue
                            if  run_name is not None and run_name.startswith(desired_date):
                                return True

    elif file_name.lower().endswith(".xlsx"):
        # Open the Excel file
        try:
            workbook = openpyxl.load_workbook(filename=report_path, data_only=True)
        except:
            return False
        # Iterate through sheets
        for excel_sheet_name in workbook.sheetnames:
            sheet = workbook[excel_sheet_name]
            for i in range(1,sheet.max_row):
                for j in range(1, sheet.max_column):
                    cell_value = sheet.cell(row=i, column=j).value
                    if type(cell_value) is str:
                        if cell_value == "Run name":
                            try:
                                run_name = sheet.cell(row=i, column=j+1).value
                            except:
                                continue
                            if  run_name is not None and run_name.startswith(desired_date):
                                return True
    
    return False

def sciebo_fastq_match(flowcell_id, sciebo_file):
    report_path = sciebo_file

    if sciebo_file.lower().endswith(".xls"):
        # Open the Excel file
        workbook = xlrd.open_workbook(report_path)
        # Iterate through sheets
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            # Iterate through rows and columns
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(i, j)
                    if type(cell_value) is str:
                        result_list = re.split(', | ', cell_value)
                        threshold = 1
                        for word in result_list:
                            distance = Levenshtein.distance(flowcell_id, word)
                            if distance <= threshold:
                                return True
                       
    elif sciebo_file.lower().endswith(".xlsx"):
        # Open the Excel file
        workbook = openpyxl.load_workbook(filename=report_path, data_only=True)
        # Iterate through sheets
        for excel_sheet_name in workbook.sheetnames:
            sheet = workbook[excel_sheet_name]
            for i in range(1,sheet.max_row):
                for j in range(1, sheet.max_column):
                    cell_value = sheet.cell(row=i, column=j).value
                    if type(cell_value) is str:
                        result_list = re.split(', | ', cell_value)
                        threshold = 1
                        for word in result_list:
                            distance = Levenshtein.distance(flowcell_id, word)
                            if distance <= threshold:
                                return True              

    return False

def has_three_months_passed(date_str):
    # Extract date part (assuming the format is YYMMDD at the beginning of the string)
    date_part = date_str[:6]
    # Convert the date part to a datetime object
    date_format = '%y%m%d'
    extracted_date = datetime.strptime(date_part, date_format)
    # Calculate the date 3 months from the extracted date
    three_months_later = extracted_date + timedelta(days=90)
    # Get the current date
    current_date = datetime.now()

    # Check if the current date is greater than or equal to the date 3 months from the extracted date
    return current_date >= three_months_later