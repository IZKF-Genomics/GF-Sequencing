import os
import re
import openpyxl
import xlrd
import Levenshtein
import pandas as pd

from datetime import datetime
from fastq_parser import parse_fastq_stats_folder, parse_multiqc_files
from sciebo_parser import parse_sciebo_report

# Predefined dictionary for mapping sequencing kits to expected clusters
sequencing_kit_to_clusters = {
    'nextseq 500/550 high output kit v2.5 (75 cycles)': '400 mio.',
    'nextseq 500/550 high output kit v2.5 (150 cycles)': '400 mio.',
    'nextseq 500/550 mid output kit v2.5 (300 cycles)': '130 mio.',
    'nextseq 500/550 mid output kit v2.5 (150 cycles)': '130 mio.',
    'miseq reagent kit v3 (150-cycle)': '22–25 million',
    'miseq reagent kit v3 (600-cycles)': '22–25 million',
    'miseq reagent kit v2 (50-cycles)': '12-15 million',
    'miseq reagent kit v2 (300-cycles)': '12-15 million',
    'miseq reagent kit v2 (500-cycles)': '12-15 million',
    'miseq reagent micro kit v2 (300-cycles)': '4 million',
    'miseq reagent nano kit v2 (300-cycles)': '1 million',
    'miseq reagent nano kit v2 (500-cycles)': '1 million',
    'novaseq 6000 s4 reagent kit v1.5 (300 cycles)': '8–10 billion',
    'novaseq 6000 s4 reagent kit v1.5 (200 cycles)': '8–10 billion',
    'novaseq 6000 s4 reagent kit v1.5 (35 cycles)': '8–10 billion',
    'novaseq 6000 s2 reagent kit v1.5 (300 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s2 reagent kit v1.5 (200 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s2 reagent kit v1.5 (100 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s1 reagent kit v1.5 (300 cycles)': '1.3–1.6 billion',
    'novaseq 6000 s1 reagent kit v1.5 (200 cycles)': '1.3–1.6 billion',
    'novaseq 6000 s1 reagent kit v1.5 (100 cycles)': '1.3–1.6 billion',
    'novaseq 6000 sp 500 cycles': '650–800 million',
    'novaseq 6000 sp 300 cycles': '650–800 million',
    'novaseq 6000 sp 200 cycles)': '650–800 million',
    'novaseq 6000 sp 100 cycles': '650–800 million',
    }

sequencing_kit_to_max_clusters = {
    'nextseq 500/550 high output kit v2.5 (75 cycles)': 400000000,
    'nextseq 500/550 high output kit v2.5 (150 cycles)': 400000000,
    'nextseq 500/550 mid output kit v2.5 (300 cycles)': 130000000,
    'nextseq 500/550 mid output kit v2.5 (150 cycles)': 130000000,
    'miseq reagent kit v3 (150-cycle)': 25000000,
    'miseq reagent kit v3 (600-cycles)': 25000000,
    'miseq reagent kit v2 (50-cycles)': 15000000,
    'miseq reagent kit v2 (300-cycles)': 15000000,
    'miseq reagent kit v2 (500-cycles)': 15000000,
    'miseq reagent micro kit v2 (300-cycles)': 4000000,
    'miseq reagent nano kit v2 (300-cycles)': 1000000,
    'miseq reagent nano kit v2 (500-cycles)': 1000000,
    'novaseq 6000 s4 reagent kit v1.5 (300 cycles)': 10000000000,
    'novaseq 6000 s4 reagent kit v1.5 (200 cycles)': 10000000000,
    'novaseq 6000 s4 reagent kit v1.5 (35 cycles)': 10000000000,
    'novaseq 6000 s2 reagent kit v1.5 (300 cycles)': 4100000000,
    'novaseq 6000 s2 reagent kit v1.5 (200 cycles)': 4100000000,
    'novaseq 6000 s2 reagent kit v1.5 (100 cycles)': 4100000000,
    'novaseq 6000 s1 reagent kit v1.5 (300 cycles)': 1600000000,
    'novaseq 6000 s1 reagent kit v1.5 (200 cycles)': 1600000000,
    'novaseq 6000 s1 reagent kit v1.5 (100 cycles)': 1600000000,
    'novaseq 6000 sp 500 cycles': 800000000, 
    'novaseq 6000 sp 300 cycles': 800000000,
    'novaseq 6000 sp 200 cycles)': 800000000,
    'novaseq 6000 sp 100 cycles': 800000000,
    }

expected_reading_per_sample_mapping = {
    'RNAseq': None,
    'tRNAseq': 50000000,
    'mRNAseq': 30000000, 
    '3mRNAseq': 10000000, 
    'ChIPseq': None, 
    'ATACseq': 50000000, 
    'ampliseq': None, 
    'scRNAseq': None,    
    'scVDJseq': None, 
    'scATACseq': None, 
    'miRNAseq': 10000000, 
    'BWGS': None, 
    'WES': None, 
    'fastq': None, 
    '16S': None, 
    'MAG': None
    }

fastq_folder_path = os.path.join("..", "fastq")
bcl_folder_path = os.path.join("..", "fastq")
sciebo_folder_path = os.path.join("..", "statistics","sciebo")


def main():
    fastq_folders = os.listdir(fastq_folder_path)
    fastq_dates = get_fastq_date(fastq_folders)
    fastq_sequencers = get_sequencer(fastq_folders)
    data = {
        "Project Name": fastq_folders,
        "Protocol Name":[None] * len(fastq_folders),
        "Date": fastq_dates,
        "Sciebo Found": [None] * len(fastq_folders),
        "Application":[None] * len(fastq_folders),
        "Sequencer": fastq_sequencers,
        "STD in Millions": [None] * len(fastq_folders),
        "CV": [None] * len(fastq_folders),
        "Undetermined Reads Percentage": [None] * len(fastq_folders),
        "Most Common Undetermined Barcode": [None] * len(fastq_folders),
        "Most Common Undetermined Barcode Percentage": [None] * len(fastq_folders),
        "Undetermined Distribution String": [None] * len(fastq_folders),
        "Read Distribution": [None] * len(fastq_folders),
        "Expected Reading Per Sample": [None] * len(fastq_folders),
        "Number of Samples Above Requirement": [None] * len(fastq_folders),
        "Number of Samples Below Requirement": [None] * len(fastq_folders),
        "Sequencing Kit": [None] * len(fastq_folders),
        "Cycles Read 1": [None] * len(fastq_folders),
        "Cycles Index 1": [None] * len(fastq_folders),
        "Cycles Read 2": [None] * len(fastq_folders),
        "Cycles Index 2": [None] * len(fastq_folders),
        "Density": [None] * len(fastq_folders),
        "Clusters PF": [None] * len(fastq_folders),
        "Yields": [None] * len(fastq_folders),
        "Q 30": [None] * len(fastq_folders),
        "Phix Input Percent": [None] * len(fastq_folders),
        "Phix Output Percent": [None] * len(fastq_folders),
        "Phix Barcode": [None] * len(fastq_folders),
        "Name": [None] * len(fastq_folders),
        "Total Read Count in Millions": [None] * len(fastq_folders),
        "Max Cluster": [None] * len(fastq_folders),
    }
    
    df = pd.DataFrame(data)
    df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y')
    df = df.sort_values(by="Date")
    df = df.set_index(['Project Name'])

    for folder in fastq_folders:
        # First check that it is a sequencing folder
        if re.match(r'^\d{6}', folder) is None:
            print("The folder is not in the expected fastq project format")
            continue
        df.loc[folder,'STD in Millions'] , df.loc[folder,'CV'], df.loc[folder,'Undetermined Reads Percentage'], df.loc[folder,'Read Distribution'], df.loc[folder,'Total Read Count in Millions']  = parse_multiqc_files(folder)
        df.loc[folder,'Most Common Undetermined Barcode'], df.loc[folder,'Undetermined Distribution String'], df.loc[folder,'Most Common Undetermined Barcode Percentage'], df.loc[folder,'Phix Output Percent'], df.loc[folder,'Phix Barcode'] =  parse_fastq_stats_folder(folder)
        sciebo_report_path = find_corresponding_sciebo(folder)
        if sciebo_report_path is not None:
            # sciebo_report_path = os.path.join(sciebo_folder_path, sciebo_report_path)
            df.loc[folder,'Sequencing Kit'], df.loc[folder,'Cycles Read 1'], df.loc[folder,'Cycles Index 1'], df.loc[folder,'Cycles Read 2'], df.loc[folder,'Cycles Index 2'], df.loc[folder,'Density'], df.loc[folder,'Clusters PF'], df.loc[folder,'Yields'], df.loc[folder,'Q 30'], df.loc[folder,'Name'], df.loc[folder, 'Protocol Name'], df.loc[folder, 'Application'], df.loc[folder, 'Phix Input Percent'] = parse_sciebo_report(sciebo_report_path) 
            df.loc[folder,"Sciebo Found"] = True
        else:
            df.loc[folder,"Sciebo Found"] = False    
            
    df['Sequencing Kit'] = df['Sequencing Kit'].str.lower()
    df['Expected Clusters'] = df['Sequencing Kit'].map(sequencing_kit_to_clusters)
    # df['maximal expected clusters'] = df['sequencing_kit'].map(sequencing_kit_to_max_clusters)
    df['Ratio Total Read Count and Expected Cluster'] = pd.to_numeric(df['Total Read Count in Millions'], errors='coerce') / df['Sequencing Kit'].map(sequencing_kit_to_max_clusters)
    df['Ratio Total Read Count and Expected Cluster'] = df['Ratio Total Read Count and Expected Cluster'].round(2)


    # Calculate the phix output percentage
    df['Phix Output Percent'] = pd.to_numeric(df['Phix Output Percent'], errors='coerce') / pd.to_numeric(df['Total Read Count in Millions'], errors='coerce') 
    
    # Adjust the scaling of the columns:
    df["Phix Input Percent"] = pd.to_numeric(df["Phix Input Percent"], errors='coerce')

    df['Phix Output Percent'] *= 100
    df['Phix Input Percent'] *= 100
    df["Phix Output Percent"] = df["Phix Output Percent"].round(2)
    df["Phix Input Percent"] = df["Phix Input Percent"].round(2)

    df["Total Read Count in Millions"] = pd.to_numeric(df["Total Read Count in Millions"], errors='coerce')
    df["Total Read Count in Millions"] /= 1000000
    df["Total Read Count in Millions"] = df["Total Read Count in Millions"].round(2)

    # calculate the number of samples meeting the reads requirement
    df['Expected Reading Per Sample'] = df['Application'].map(expected_reading_per_sample_mapping)
    df['Number of Samples Above Requirement'], df['Number of Samples Below Requirement'] = zip(*df.apply(check_read_count, axis=1))

    df.drop("Max Cluster", axis=1, inplace=True)
    df.drop("Expected Reading Per Sample", axis=1, inplace=True)

    return df


def check_read_count(row):
    """ Count above and below requirement samples"""
    samples_above_requirement = 0
    samples_below_requirement = 0
    expected_reading_per_sample = row['Expected Reading Per Sample']
    total_read_count = row["Total Read Count in Millions"] * 1000000
    sample_read_distribution = row['Read Distribution']

    if expected_reading_per_sample is None or sample_read_distribution is None:
        return None, None  # No comparison for 'None' types
    
    sample_read_distribution = [float(x) for x in sample_read_distribution.split('-')]
    # Convert from percenatges to actuall read numbers
    sample_read_distribution = [x * total_read_count for x in sample_read_distribution]
    # Drop undetemined reads
    sample_read_distribution = sample_read_distribution[:-1]

    for sample_reads_count in sample_read_distribution:
        if sample_reads_count < expected_reading_per_sample:
            # Number of Samples Above Requirement increased by 1
            samples_below_requirement += 1  
        elif sample_reads_count >= expected_reading_per_sample:
            # Number of Samples Below Requirement increased by 1
            samples_above_requirement += 1 

    return samples_above_requirement, samples_below_requirement  # No change in counts


def find_corresponding_sciebo(fastq_folder):
    sequence_date_prefix = fastq_folder.split('_')[0]
    flowcell_id = fastq_folder.split('_')[3][1:]

    # sciebo_files = os.listdir(sciebo_folder_path) # Was used previously when all sciebo files were in one directory
    sciebo_files = []
    # Walk through the directory tree
    for folder_path, _, files in os.walk(sciebo_folder_path):
        for file_name in files:
            # Construct the full path to the  sciebo file
            sciebo_file_path = os.path.join(folder_path, file_name)
            sciebo_files.append(sciebo_file_path)
    sciebo_candidates = [file for file in sciebo_files if sciebo_date_match(file, sequence_date_prefix)] # os.path.basename(file).startswith(sequence_date_prefix)]

    if len(sciebo_candidates) == 0:
        print(f"zero sciebo candidates")
        return None
    elif len(sciebo_candidates) == 1:
        print(f"single sciebo match found! {sciebo_candidates[0]}")
        return sciebo_candidates[0]
    else:
        for sciebo_file in sciebo_candidates: 
            if sciebo_fastq_match(flowcell_id, sciebo_file):
                print(f"sciebo match found from multiple candidates! {sciebo_file}")
                return sciebo_file
    print(f"No sciebo match was found for {fastq_folder}")
    return None

def sciebo_date_match(file_name, desired_date):
    """ Find if the sciebo corresponds to the specific run date
    """
    # print(f"Does sciebo: '{file_name}' corresponds to the specific run date:'{desired_date}'")
    report_path = file_name
    if file_name.lower().endswith(".xls"):
        # Open the Excel file
        try:
            workbook = xlrd.open_workbook(report_path)
        except:
            return False
        # Iterate through sheets
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            # Iterate through rows and columns
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(i, j)
                    if type(cell_value) is str:
                        if cell_value == "Run name":
                            try:
                                run_name = sheet.cell_value(i, j + 1)
                            except:
                                continue
                            if  run_name is not None and run_name.startswith(desired_date):
                                return True

    elif file_name.lower().endswith(".xlsx"):
        # Open the Excel file
        try:
            workbook = openpyxl.load_workbook(filename=report_path, data_only=True)
        except:
            return False
        # Iterate through sheets
        for excel_sheet_name in workbook.sheetnames:
            sheet = workbook[excel_sheet_name]
            for i in range(1,sheet.max_row):
                for j in range(1, sheet.max_column):
                    cell_value = sheet.cell(row=i, column=j).value
                    if type(cell_value) is str:
                        if cell_value == "Run name":
                            try:
                                run_name = sheet.cell(row=i, column=j+1).value
                            except:
                                continue
                            if  run_name is not None and run_name.startswith(desired_date):
                                return True
    return False

def extract_date_from_folder(folder_name):
    try:
        # Extracting the YYMMDD part from the folder name
        date_str = folder_name[:6]
        # Converting the date string to a datetime object
        date_obj = datetime.strptime(date_str, '%y%m%d')
        # Formatting the datetime object as 'DD.MM.YYYY'
        formatted_date = date_obj.strftime('%d.%m.%Y')
        return formatted_date

    except ValueError:
        # Handle the case where the folder name doesn't match the expected format
        return None


def get_fastq_date(fastq_folders):
    dates = [extract_date_from_folder(folder) for folder in fastq_folders]
    return dates


def get_sequencer(fastq_folders):
    sequencers = [extract_sequencer_from_folder(folder) for folder in fastq_folders]
    return sequencers


def extract_sequencer_from_folder(folder_name):
    name_array = folder_name.split('_')
    if len(name_array) <= 1 :
        return ""
    instrument_id = name_array[1]
    if instrument_id.startswith("NB501289"):
        return "nextseq500"
    elif instrument_id.startswith("M00818"):
        return "miseq1"
    elif instrument_id.startswith("M04404"):
        return "miseq2"
    elif instrument_id.startswith("A01742"):
        return "novaseq"
    else: 
        return ""


def sciebo_fastq_match(flowcell_id, sciebo_file):
    report_path = sciebo_file

    if sciebo_file.lower().endswith(".xls"):
        # Open the Excel file
        workbook = xlrd.open_workbook(report_path)
        # Iterate through sheets
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            # Iterate through rows and columns
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(i, j)
                    if type(cell_value) is str:
                        result_list = re.split(', | ', cell_value)
                        threshold = 1
                        for word in result_list:
                            distance = Levenshtein.distance(flowcell_id, word)
                            if distance <= threshold:
                                return True
                       

    elif sciebo_file.lower().endswith(".xlsx"):
        # Open the Excel file
        workbook = openpyxl.load_workbook(filename=report_path, data_only=True)
        # Iterate through sheets
        for excel_sheet_name in workbook.sheetnames:
            sheet = workbook[excel_sheet_name]
            for i in range(1,sheet.max_row):
                for j in range(1, sheet.max_column):
                    cell_value = sheet.cell(row=i, column=j).value
                    if type(cell_value) is str:
                        result_list = re.split(', | ', cell_value)
                        threshold = 1
                        for word in result_list:
                            distance = Levenshtein.distance(flowcell_id, word)
                            if distance <= threshold:
                                return True              
    return False


def format_with_commas(number):
    if pd.notna(number):  # Check if the value is not None
        return '{:,}'.format(number)
    else:
        return None


if __name__ == '__main__':
    """
    Create a data frame containing different statistics from all the sequencing projects
    """
    df = main()
    df.to_csv('sequencing_statistics.csv', index=True)

